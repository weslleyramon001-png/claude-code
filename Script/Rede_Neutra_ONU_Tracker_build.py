import zipfile, io, os
import xml.etree.ElementTree as ET
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList
from openpyxl.utils import get_column_letter as L

BG="060E1C"; BG_CARD="0C1A2E"; BG_CARD2="111F36"; BG_HDR="071525"
EMERALD="10B981"; EM_D="059669"; EM_L="34D399"; EM_DIM="052E16"
GOLD="F59E0B"; RED="F87171"; BLUE_L="93C5FD"; GRAY="94A3B8"; WHITE="F0F9FF"

def F(c):  return PatternFill("solid", fgColor=c)
def fn(c,s=10,bold=False): return Font(name="Calibri",color=c,size=s,bold=bold)
CA = Alignment(horizontal="center",vertical="center",wrap_text=True)
def Bdr(left=None,right=None,top=None,bottom=None):
    def s(t): return Side(border_style=t[1],color=t[0]) if t else Side(border_style=None)
    return Border(left=s(left),right=s(right),top=s(top),bottom=s(bottom))
def fill_row(ws,row,cols,color):
    for c in range(1,cols+1):
        try: ws.cell(row=row,column=c).fill=F(color)
        except: pass

CARDS=[(2,5),(7,10),(12,15),(17,20)]; TOTAL_COLS=21

def set_col_widths(ws):
    pattern=[2,12,12,12,12,2,12,12,12,12,2,12,12,12,12,2,12,12,12,12,2]
    for i,w in enumerate(pattern,1): ws.column_dimensions[L(i)].width=w
    for i in range(22,35): ws.column_dimensions[L(i)].width=12

def kpi_card(ws,ci,rl,rv,label,value,bg=BG_CARD,lc=GRAY,vc=EMERALD,vs=28,ac=EMERALD):
    cs,ce=CARDS[ci]
    ws.merge_cells(f"{L(cs)}{rl}:{L(ce)}{rl}")
    c=ws.cell(row=rl,column=cs)
    c.value=label; c.font=fn(lc,9,True); c.fill=F(bg); c.alignment=CA
    c.border=Bdr(left=(ac,"medium"),right=(ac,"medium"),top=(ac,"medium"))
    ws.merge_cells(f"{L(cs)}{rv}:{L(ce)}{rv}")
    c=ws.cell(row=rv,column=cs)
    c.value=value; c.font=Font(name="Calibri",size=vs,bold=True,color=vc)
    c.fill=F(bg); c.alignment=CA
    c.border=Bdr(left=(ac,"medium"),right=(ac,"medium"),bottom=(ac,"medium"))

OPERADORAS=[
    ("VIRTUAL NET",      3,8, 720.00, 90.00,"0%",   "✅ OK"),
    ("INEDITTUS FIBRA",  2,5, 499.50, 99.90,"0%",   "✅ OK"),
    ("POPNET",           2,4, 360.00, 90.00,"25%",  "⚠️ Atenção"),
    ("RSNET",            1,3, 267.00, 89.00,"0%",   "✅ OK"),
    ("WORLD CONNECT",    2,5, 449.50, 89.90,"20%",  "⚠️ Atenção"),
]
TOT_CTO=sum(o[1] for o in OPERADORAS)
TOT_ONU=sum(o[2] for o in OPERADORAS)
TOT_REC=sum(o[3] for o in OPERADORAS)

def build_dashboard(wb):
    ws=wb.create_sheet("📊 Dashboard")
    ws.sheet_view.showGridLines=False; ws.sheet_view.showRowColHeaders=False
    for r in range(1,55): ws.row_dimensions[r].height=22
    set_col_widths(ws)
    for r in range(1,55): fill_row(ws,r,TOTAL_COLS+12,BG)
    ws.row_dimensions[2].height=10; ws.row_dimensions[3].height=50
    ws.row_dimensions[4].height=24; ws.row_dimensions[5].height=10
    ws.merge_cells(f"B3:{L(TOTAL_COLS)}3")
    c=ws.cell(row=3,column=2)
    c.value="🌐  REDE NEUTRA ONU TRACKER  —  Gestão Integrada de Operadoras"
    c.font=Font(name="Calibri",size=22,bold=True,color=EMERALD)
    c.fill=F(BG_HDR); c.alignment=CA
    ws.merge_cells(f"B4:{L(TOTAL_COLS)}4")
    c=ws.cell(row=4,column=2)
    c.value=f"Operadoras Ativas: {len(OPERADORAS)}  |  CTOs: {TOT_CTO}  |  ONUs: {TOT_ONU}  |  Receita Mensal: R$ {TOT_REC:,.2f}".replace(",",".")
    c.font=fn(GRAY,11); c.fill=F(BG_HDR); c.alignment=CA
    ws.row_dimensions[6].height=20; ws.row_dimensions[7].height=46; ws.row_dimensions[8].height=10
    kpi_card(ws,0,6,7,"🏗️  CTOs Cadastradas",TOT_CTO,vs=36)
    kpi_card(ws,1,6,7,"✅  CTOs Ativas",TOT_CTO-1,vs=36,vc=EM_L,ac=EM_L)
    kpi_card(ws,2,6,7,"📡  ONUs Conectadas",TOT_ONU,vs=36,vc=BLUE_L,ac=BLUE_L)
    kpi_card(ws,3,6,7,"🏢  Operadoras Ativas",len(OPERADORAS),vs=36,vc=GOLD,ac=GOLD)
    ws.row_dimensions[9].height=20; ws.row_dimensions[10].height=46; ws.row_dimensions[11].height=10
    kpi_card(ws,0,9,10,"💰  Receita Mensal",f"R$ {TOT_REC:,.2f}".replace(",","."),vs=24,vc=GOLD,ac=GOLD)
    kpi_card(ws,1,9,10,"📈  Ticket Médio/ONU",f"R$ {TOT_REC/TOT_ONU:.2f}".replace(".",","),vs=24,vc=EM_L,ac=EM_L)
    kpi_card(ws,2,9,10,"⚠️  Operadoras c/ Alerta","2",vs=36,vc=RED,ac=RED,bg=BG_CARD2)
    kpi_card(ws,3,9,10,"🔋  Uptime Médio Rede","95%",vs=36,vc=EMERALD,ac=EMERALD)
    ws.row_dimensions[12].height=8; ws.row_dimensions[13].height=26
    for r in [14,15,16,17,18]: ws.row_dimensions[r].height=22
    ws.merge_cells(f"B13:{L(TOTAL_COLS)}13")
    c=ws.cell(row=13,column=2)
    c.value="▶  RESUMO POR OPERADORA"; c.font=fn(EMERALD,12,True)
    c.fill=F(BG_HDR); c.alignment=Alignment(horizontal="left",vertical="center")
    HDR_COLS=["Operadora","CTOs","ONUs","Receita/mês","Ticket Médio","Inadimp.","Status"]
    tbl_cols=[2,5,7,9,12,15,17]; tbl_end=[4,6,8,11,14,16,20]
    for h,cs,ce in zip(HDR_COLS,tbl_cols,tbl_end):
        ws.merge_cells(f"{L(cs)}14:{L(ce)}14")
        c=ws.cell(row=14,column=cs)
        c.value=h; c.font=fn(EMERALD,9,True); c.fill=F(EM_DIM); c.alignment=CA
        c.border=Border(left=Side("thin",color=EMERALD),right=Side("thin",color=EMERALD),
                        top=Side("thin",color=EMERALD),bottom=Side("thin",color=EMERALD))
    for ri,(op,cto,onu,rec,tick,inadimp,status) in enumerate(OPERADORAS):
        row=15+ri
        data=[op,cto,onu,f"R$ {rec:,.2f}".replace(",","."),
              f"R$ {tick:.2f}".replace(".",","),inadimp,status]
        vc2=RED if "Atenção" in status else EM_L
        for i,(val,cs,ce) in enumerate(zip(data,tbl_cols,tbl_end)):
            ws.merge_cells(f"{L(cs)}{row}:{L(ce)}{row}")
            c=ws.cell(row=row,column=cs); c.value=val
            col=GOLD if i==0 else (vc2 if i==6 else WHITE)
            c.font=fn(col,9,i==0)
            c.fill=F(BG_CARD if ri%2==0 else BG_CARD2); c.alignment=CA
            c.border=Border(left=Side("thin",color=BG_HDR),right=Side("thin",color=BG_HDR),
                            bottom=Side("thin",color=BG_HDR))
    for r in range(19,50): ws.row_dimensions[r].height=22
    DATA_ROW=2
    ws.cell(row=DATA_ROW,column=23).value="Operadora"
    ws.cell(row=DATA_ROW,column=24).value="ONUs"
    ws.cell(row=DATA_ROW,column=25).value="Receita"
    for i,(op,cto,onu,rec,*_) in enumerate(OPERADORAS):
        r=DATA_ROW+1+i
        ws.cell(row=r,column=23).value=op; ws.cell(row=r,column=24).value=onu; ws.cell(row=r,column=25).value=rec
    ch1=BarChart(); ch1.type="col"; ch1.grouping="clustered"
    ch1.title="ONUs por Operadora"; ch1.style=10; ch1.width=17; ch1.height=13
    cats1=Reference(ws,min_col=23,min_row=DATA_ROW+1,max_row=DATA_ROW+5)
    vals1=Reference(ws,min_col=24,min_row=DATA_ROW,max_row=DATA_ROW+5)
    ch1.add_data(vals1,titles_from_data=True); ch1.set_categories(cats1)
    ch1.series[0].graphicalProperties.solidFill=EMERALD
    ws.add_chart(ch1,"B20")
    ch2=PieChart(); ch2.title="Receita por Operadora"; ch2.style=10
    ch2.width=17; ch2.height=13
    vals2=Reference(ws,min_col=25,min_row=DATA_ROW,max_row=DATA_ROW+5)
    cats2=Reference(ws,min_col=23,min_row=DATA_ROW+1,max_row=DATA_ROW+5)
    ch2.add_data(vals2,titles_from_data=True); ch2.set_categories(cats2)
    dl=DataLabelList(); dl.showPercent=True; ch2.dLbls=dl
    for idx,color in enumerate([EMERALD,GOLD,BLUE_L,RED,"A78BFA"]):
        pt=DataPoint(idx=idx); pt.graphicalProperties.solidFill=color
        ch2.series[0].dPt.append(pt)
    ws.add_chart(ch2,"L20")

SHEET_ICONS=["🔵","🟢","🟡","🔴","🟣"]

def build_operator_sheet(wb,op_data,icon):
    op,cto_n,onu_n,rec,tick,inadimp,status=op_data
    ws=wb.create_sheet(f"{icon} {op[:12]}")
    ws.sheet_view.showGridLines=False; ws.sheet_view.showRowColHeaders=False
    set_col_widths(ws)
    for r in range(1,45): ws.row_dimensions[r].height=22
    for r in range(1,45): fill_row(ws,r,TOTAL_COLS,BG)
    ws.row_dimensions[3].height=46
    ws.merge_cells(f"B3:{L(TOTAL_COLS)}3")
    c=ws.cell(row=3,column=2)
    c.value=f"{icon}  {op}  —  Gestão de CTOs & ONUs"
    c.font=Font(name="Calibri",size=18,bold=True,color=EMERALD)
    c.fill=F(BG_HDR); c.alignment=CA
    ws.row_dimensions[4].height=22
    ws.merge_cells(f"B4:{L(TOTAL_COLS)}4")
    c=ws.cell(row=4,column=2)
    c.value=f"CTOs: {cto_n}  |  ONUs: {onu_n}  |  Receita: R$ {rec:,.2f}  |  Status: {status}".replace(",",".")
    c.font=fn(GRAY,10); c.fill=F(BG_HDR); c.alignment=CA
    ws.row_dimensions[6].height=20; ws.row_dimensions[7].height=46; ws.row_dimensions[8].height=10
    kpi_card(ws,0,6,7,"🏗️  CTOs",cto_n,vs=36)
    kpi_card(ws,1,6,7,"📡  ONUs",onu_n,vs=36,vc=BLUE_L,ac=BLUE_L)
    kpi_card(ws,2,6,7,"💰  Receita/mês",f"R$ {rec:,.2f}".replace(",","."),vs=22,vc=GOLD,ac=GOLD)
    kpi_card(ws,3,6,7,"📊  Inadimplência",inadimp,vs=36,
             vc=RED if "%" in inadimp and inadimp!="0%" else EMERALD,
             ac=RED if "%" in inadimp and inadimp!="0%" else EMERALD)
    ws.row_dimensions[10].height=26
    HDR=["CTO","Endereço","ONUs Ativas","ONUs Total","Manutenção","Status"]
    tcols=[2,4,8,11,14,17]; tends=[3,7,10,13,16,20]
    for h,cs,ce in zip(HDR,tcols,tends):
        ws.merge_cells(f"{L(cs)}10:{L(ce)}10")
        c=ws.cell(row=10,column=cs)
        c.value=h; c.font=fn(EMERALD,9,True); c.fill=F(EM_DIM); c.alignment=CA
        c.border=Border(left=Side("thin",color=EMERALD),right=Side("thin",color=EMERALD),
                        top=Side("thin",color=EMERALD),bottom=Side("thin",color=EMERALD))
    sample=[("CTO-01","Rua das Flores, 123",onu_n//cto_n+1,onu_n//cto_n+2,"—","✅ OK"),
            ("CTO-02","Av. Principal, 456",onu_n//cto_n,onu_n//cto_n+1,"Ago/26","✅ OK")]
    if cto_n>2: sample.append(("CTO-03","Rua Nova, 789",onu_n//cto_n-1,onu_n//cto_n,"—","⚠️ Verificar"))
    for ri,(cto_id,addr,oactive,otot,manut,st) in enumerate(sample):
        row=11+ri; data=[cto_id,addr,oactive,otot,manut,st]; vc2=RED if "Verificar" in st else EM_L
        for val,cs,ce in zip(data,tcols,tends):
            ws.merge_cells(f"{L(cs)}{row}:{L(ce)}{row}")
            c=ws.cell(row=row,column=cs); c.value=val
            col=GOLD if val==cto_id else (vc2 if val==st else WHITE)
            c.font=fn(col,9); c.fill=F(BG_CARD if ri%2==0 else BG_CARD2); c.alignment=CA
            c.border=Border(left=Side("thin",color=BG_HDR),right=Side("thin",color=BG_HDR),
                            bottom=Side("thin",color=BG_HDR))

# ══════════════════════════════════════════════════════════════════════════
# ET-based patch — exactly like V4 (which worked), adds txPr in correct positions
# ══════════════════════════════════════════════════════════════════════════
C_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
A_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
C_ = f"{{{C_NS}}}"; A_ = f"{{{A_NS}}}"

def mk_spPr(fill_hex, ln_hex):
    sp = ET.Element(f"{C_}spPr")
    sf = ET.SubElement(ET.SubElement(sp, f"{A_}solidFill"), f"{A_}srgbClr"); sf.set("val", fill_hex)
    ln = ET.SubElement(sp, f"{A_}ln")
    lf = ET.SubElement(ET.SubElement(ln, f"{A_}solidFill"), f"{A_}srgbClr"); lf.set("val", ln_hex)
    return sp

def mk_txPr(fill_hex, bold=False):
    txPr = ET.Element(f"{C_}txPr")
    ET.SubElement(txPr, f"{A_}bodyPr")
    lst = ET.SubElement(txPr, f"{A_}lstStyle")
    defPPr = ET.SubElement(lst, f"{A_}defPPr")
    defRPr = ET.SubElement(defPPr, f"{A_}defRPr")
    if bold: defRPr.set("b", "1")
    sf = ET.SubElement(ET.SubElement(defRPr, f"{A_}solidFill"), f"{A_}srgbClr"); sf.set("val", fill_hex)
    lat = ET.SubElement(defRPr, f"{A_}latin"); lat.set("typeface", "Calibri")
    return txPr

def insert_before(parent, tag, new_elem):
    """Insert new_elem before the first child with given tag."""
    children = list(parent)
    for i, child in enumerate(children):
        if child.tag == tag:
            parent.insert(i, new_elem)
            return
    parent.append(new_elem)  # fallback

def patch_chart_xml(data: bytes) -> bytes:
    # Register namespaces so ET serializes with c: and a: prefixes (like V4)
    ns_map = {}
    for event, elem in ET.iterparse(io.BytesIO(data), events=["start-ns"]):
        ns_map[elem[0]] = elem[1]
    for prefix, uri in ns_map.items():
        try: ET.register_namespace(prefix, uri)
        except: pass
    ET.register_namespace("c", C_NS)
    ET.register_namespace("a", A_NS)

    tree = ET.fromstring(data)
    chart_el = tree.find(f"{C_}chart")

    # ── 1. chartSpace spPr → append at end (after </chart>) ──────────────
    tree.append(mk_spPr("0A1628", "10B981"))

    if chart_el is not None:
        pa = chart_el.find(f"{C_}plotArea")
        if pa is not None:
            # ── 2. plotArea spPr → append at end ─────────────────────────
            pa.append(mk_spPr("060E1C", "059669"))

            # ── 3. catAx txPr → insert before <crossAx> ──────────────────
            for ax in pa.findall(f"{C_}catAx"):
                insert_before(ax, f"{C_}crossAx", mk_txPr("F0F9FF"))

            # ── 4. valAx txPr → insert before <crossAx> ──────────────────
            for ax in pa.findall(f"{C_}valAx"):
                insert_before(ax, f"{C_}crossAx", mk_txPr("F0F9FF"))

        # ── 5. title txPr → append at end (after </tx>) ──────────────────
        title_el = chart_el.find(f"{C_}title")
        if title_el is not None:
            title_el.append(mk_txPr("10B981", bold=True))

        # ── 6. legend spPr + txPr → append at end ────────────────────────
        legend = chart_el.find(f"{C_}legend")
        if legend is not None:
            legend.append(mk_spPr("0C1A2E", "10B981"))
            legend.append(mk_txPr("F0F9FF"))

    # ── 7. dLbls txPr → append at end ────────────────────────────────────
    for dLbls in tree.iter(f"{C_}dLbls"):
        dLbls.append(mk_txPr("F0F9FF"))

    return ET.tostring(tree, encoding="unicode").encode("utf-8")

def patch_xlsx_charts(path_in, path_out):
    with zipfile.ZipFile(path_in,"r") as zin:
        with zipfile.ZipFile(path_out,"w",zipfile.ZIP_DEFLATED) as zout:
            for name in zin.namelist():
                data = zin.read(name)
                if name.startswith("xl/charts/chart") and name.endswith(".xml"):
                    try:
                        data = patch_chart_xml(data)
                        print(f"  patched {name}")
                    except Exception as e:
                        print(f"  [warn] {name}: {e}")
                zout.writestr(name, data)

# ── Build ──────────────────────────────────────────────────────────────────
wb=Workbook(); wb.remove(wb.active)
build_dashboard(wb)
for op_data,icon in zip(OPERADORAS,SHEET_ICONS):
    build_operator_sheet(wb,op_data,icon)

TMP="/tmp/Rede_Neutra_V8_raw.xlsx"
OUT="/tmp/Rede_Neutra_V8.xlsx"
wb.save(TMP); print(f"Saved: {TMP}")
patch_xlsx_charts(TMP, OUT)

# Verify namespace format matches V4
import zipfile as zf2
with zf2.ZipFile(OUT) as z:
    c = z.read("xl/charts/chart1.xml").decode()
    import re
    decl = re.search(r'<[^>]*chartSpace[^>]*>', c)
    print(f"\nNamespace decl: {decl.group()[:120]}")
    
    cat = re.search(r'<c:catAx>.*?</c:catAx>', c, re.DOTALL)
    if cat: print(f"\ncatAx:\n{cat.group()}")

print(f"\nSize: {os.path.getsize(OUT):,} bytes")
